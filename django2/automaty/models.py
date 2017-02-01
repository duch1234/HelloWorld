from __future__ import unicode_literals

from django.db import models
from django.http import HttpResponse
# Create your models here.

class dekoders(models.Model):
    dekoderName = models.CharField('data publikacji', max_length=200)
    serialNumber = models.DecimalField(max_digits=12,decimal_places=0)
    dekoderVersion = models.CharField(max_length=20)
    pub_date = models.DateTimeField('data publikacji')
  
  
  
class karta(models.Model):  
    kartaName =models.CharField('Karta', max_length=200)
    serialNumber = models.DecimalField('S/N', max_digits=12,decimal_places=0)
    kartaVersion =models.CharField('Typ', max_length=20)
    kartaOwner = models.CharField('Wlasciciel', max_length=20)
    kartaCurrentOwner= models.CharField('Obecny Wlasciciel', max_length=20)
    
    
    
    
    
class endurance(models.Model):  
    
    PENDING = 0
    DONE = 1
    STATUS_CHOICES = (
        (PENDING, 'Pending'),
        (DONE, 'Done'),
    )
    dekoderName =models.CharField('Karta', max_length=200)
    serialNumber = models.DecimalField('S/N', max_digits=12,decimal_places=0)
    RaVersion  = models.CharField('RA', max_length=1000)
    FWversion  = models.CharField('SFW', max_length=1000)
    comment =models.CharField('komentarz',  max_length=1000)
    StartDate = models.DateTimeField('poczatek')
    EndDate= models.DateTimeField('koniec',null=True, blank=True)
    statusField= models.IntegerField(choices=STATUS_CHOICES, default=0)
    


def setAsDone(modeladmin, request, queryset):
    queryset.update(statusField=1)
setAsDone.short_description = "Mark selected  as done"
#     def __unicode__(self):
#         return "{0}, {1}".format(self.dekoderName, self.serialNumber, self.dekoderVersion, self.pub_date)

def export_csv (modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    
    response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"kartaName"),
        smart_str(u"serialNumber"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.kartaName),
            smart_str(obj.serialNumber),
        ])
    return response
export_csv.short_description = u"Export CSV"





def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    #from openpyxl.cell import get_column_letter
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Report"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"kartaName", 15),
        (u"serialNumber", 15),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.kartaName,
            obj.serialNumber,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
           # c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"